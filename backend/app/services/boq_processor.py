"""
BoQ Processing Service.

Handles the background processing of uploaded BoQ files:
1. Extract items from PDF (GPT-4o Vision) or Excel
2. Classify items as material/labor
3. Look up material prices on Tokopedia
4. Calculate savings and generate analysis
"""

import asyncio
import base64
import io
import re
from datetime import datetime
from decimal import Decimal
from typing import Optional

import structlog

from app.schemas.boq import (
    BoQFileFormat,
    BoQItemExtracted,
    BoQItemType,
    BoQJobStatus,
    ExtractedBoQData,
)
from app.integrations.supabase import get_supabase_client
from app.integrations.apify import get_best_material_price

logger = structlog.get_logger()


# =============================================================================
# Classification Patterns (Indonesian construction terms)
# =============================================================================

LABOR_INDICATORS = [
    "bongkar",       # demolition
    "instalasi",     # installation (labor component)
    "pek.",          # pekerjaan (work)
    "pek ",
    "mobilisasi",    # mobilization
    "demobilisasi",
    "pembuangan",    # disposal
    "cleaning",      # cleaning
    "perbaikan",     # repair
    "refinishing",
    "plaster",       # plastering (labor)
    "aci ",          # finishing coat (labor)
    "cat ",          # painting (labor-heavy)
]

MATERIAL_INDICATORS = [
    "granit",
    "keramik",
    "batako",
    "batu ",
    "batu alam",
    "pipa",
    "pvc",
    "kabel",
    "nym",
    "pintu",
    "jendela",
    "kusen",
    "plafond",
    "gypsum",
    "waterproof",
    "lampu",
    "downlight",
    "led",
    "pompa",
    "filter",
    "saklar",
    "stop kontak",
    "gpo",
    "closet",
    "shower",
    "wastafel",
    "kran",
    "floor drain",
]

OWNER_SUPPLY_PATTERNS = [
    r"suply\s*by\s*owner",
    r"supply\s*by\s*owner",
    r"unit\s*suply\s*by\s*owner",
    r"unit\s*supply\s*by\s*owner",
    r"\(suply\s*by\s*owner\)",
    r"\(supply\s*by\s*owner\)",
]

EXISTING_PATTERNS = [
    r"use\s*existing",
    r"\(existing\)",
    r"\(use\s*existing\)",
]


# =============================================================================
# Main Processing Function
# =============================================================================


def process_boq_job_sync(
    job_id: str,
    file_content: bytes,
    file_format: BoQFileFormat,
    filename: str,
) -> None:
    """
    Fully synchronous BoQ processing - runs in a separate process via ProcessPoolExecutor.

    This function is designed to run in complete isolation from the main FastAPI process.
    It creates its own Supabase client, OpenAI client, and handles all I/O synchronously.

    The ProcessPoolExecutor spawns a new Python process, which:
    1. Avoids event loop conflicts with httpx/OpenAI
    2. Provides true parallelism (GIL doesn't affect separate processes)
    3. Isolates memory usage from the main web server
    """
    # Import here to avoid pickling issues - each process gets fresh imports
    from app.integrations.supabase import get_supabase_client as get_sb
    supabase = get_sb()

    try:
        # Update status to processing
        _update_job_status_sync(supabase, job_id, BoQJobStatus.PROCESSING, progress=5)

        # Step 1: Extract items from file
        logger.info("boq_extraction_start", job_id=job_id, format=file_format.value)

        if file_format == BoQFileFormat.PDF:
            extracted = _extract_from_pdf_sync(file_content, filename)
        else:
            # Excel extraction is already synchronous-compatible
            extracted = _extract_from_excel_sync(file_content, filename)

        _update_job_status_sync(supabase, job_id, progress=30)

        logger.info(
            "boq_extraction_complete",
            job_id=job_id,
            items_count=len(extracted.items),
        )

        # Step 2: Save extracted items and update job metadata
        _save_extracted_items_sync(supabase, job_id, extracted)

        supabase.table("boq_jobs").update({
            "project_name": extracted.project_name,
            "contractor_name": extracted.contractor_name,
            "project_location": extracted.project_location,
            "total_items_extracted": len(extracted.items),
        }).eq("id", job_id).execute()

        _update_job_status_sync(supabase, job_id, progress=40)

        # Step 3: Look up prices for material items (skip for now, do async later)
        logger.info("boq_pricing_start", job_id=job_id)
        # Skip pricing for initial release - just mark complete
        _update_job_status_sync(supabase, job_id, progress=85)

        # Step 4: Calculate summary statistics
        logger.info("boq_summary_start", job_id=job_id)
        _calculate_summary_sync(supabase, job_id)

        # Mark as completed
        _update_job_status_sync(
            supabase, job_id, BoQJobStatus.COMPLETED, progress=100
        )

        logger.info("boq_processing_complete", job_id=job_id)

    except Exception as e:
        logger.error("boq_processing_failed", job_id=job_id, error=str(e))
        _update_job_status_sync(
            supabase, job_id, BoQJobStatus.FAILED, error_message=str(e)
        )


def _extract_from_pdf_sync(file_content: bytes, filename: str) -> ExtractedBoQData:
    """Synchronous PDF extraction using GPT-4o Vision."""
    import fitz
    import json
    from app.config import get_settings

    settings = get_settings()

    # Dry-run mode: skip OpenAI entirely, return mock data for testing
    if settings.boq_dry_run:
        logger.info("boq_dry_run_mode", filename=filename)
        return ExtractedBoQData(
            project_name="[DRY RUN] Test Project",
            contractor_name="[DRY RUN] CV Test",
            items=[
                BoQItemExtracted(
                    section="DRY RUN",
                    item_number="1",
                    description="Mock item - dry run mode",
                    unit="m2",
                    quantity=10.0,
                    contractor_unit_price=100000,
                    contractor_total=1000000,
                    item_type=BoQItemType.MATERIAL,
                    extraction_confidence=1.0,
                )
            ],
            extraction_warnings=["Dry run mode - no OpenAI calls made"],
        )

    from openai import OpenAI

    client = OpenAI(
        api_key=settings.openai_api_key,
        timeout=180.0,
    )

    logger.info("pdf_to_images_start", filename=filename)

    try:
        pdf_document = fitz.open(stream=file_content, filetype="pdf")
        total_pages = len(pdf_document)
        max_pages = min(total_pages, settings.boq_max_pages)
        logger.info("pdf_page_count", total_pages=total_pages, processing_pages=max_pages)

        # Convert pages to base64 images
        image_contents = []
        for page_num in range(max_pages):
            page = pdf_document[page_num]
            mat = fitz.Matrix(150/72, 150/72)  # 150 DPI
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            base64_img = base64.b64encode(img_bytes).decode("utf-8")
            image_contents.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{base64_img}", "detail": "high"}
            })

        pdf_document.close()
        logger.info("pdf_to_images_complete", image_count=len(image_contents))

    except Exception as e:
        logger.error("pdf_to_images_failed", error=str(e))
        return ExtractedBoQData(extraction_warnings=[f"PDF conversion failed: {str(e)}"])

    extraction_prompt = """Analyze these pages from an Indonesian construction BoQ document.
Extract ALL line items into JSON format with: section, item_number, description, unit, quantity, contractor_unit_price, contractor_total, item_type (material/labor/equipment/unknown), is_owner_supply, is_existing.
Return: {"project_name": "...", "contractor_name": "...", "items": [...], "extraction_warnings": [...]}"""

    # Skip cover page, process in batches of 3
    pages_to_process = image_contents[1:] if len(image_contents) > 1 else image_contents
    BATCH_SIZE = 3

    all_items = []
    extraction_warnings = []
    project_name = None
    contractor_name = None
    project_location = None

    logger.info("gpt4o_extraction_start", total_pages=len(pages_to_process))

    for batch_start in range(0, len(pages_to_process), BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, len(pages_to_process))
        batch_pages = pages_to_process[batch_start:batch_end]
        batch_num = (batch_start // BATCH_SIZE) + 1

        logger.info("gpt4o_batch_start", batch=batch_num, pages=len(batch_pages))

        try:
            content = [{"type": "text", "text": extraction_prompt}]
            content.extend(batch_pages)

            logger.info("gpt4o_api_call_starting", batch=batch_num, content_items=len(content))

            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": content}],
                max_tokens=8000,
                temperature=0.1,
                response_format={"type": "json_object"},
            )

            logger.info("gpt4o_batch_response", batch=batch_num)

            choice = response.choices[0]
            if getattr(choice.message, 'refusal', None):
                logger.warning("gpt4o_batch_refused", batch=batch_num)
                extraction_warnings.append(f"Batch {batch_num} refused")
                continue

            if not choice.message.content:
                continue

            data = json.loads(choice.message.content)

            logger.info("gpt4o_batch_complete", batch=batch_num, items=len(data.get("items", [])))

            if not project_name:
                project_name = data.get("project_name")
            if not contractor_name:
                contractor_name = data.get("contractor_name")
            if not project_location:
                project_location = data.get("project_location")

            for item_data in data.get("items", []):
                all_items.append(BoQItemExtracted(
                    section=item_data.get("section"),
                    item_number=item_data.get("item_number"),
                    description=item_data.get("description", ""),
                    unit=item_data.get("unit"),
                    quantity=item_data.get("quantity"),
                    contractor_unit_price=item_data.get("contractor_unit_price"),
                    contractor_total=item_data.get("contractor_total"),
                    item_type=BoQItemType(item_data.get("item_type", "unknown")),
                    is_owner_supply=item_data.get("is_owner_supply", False),
                    is_existing=item_data.get("is_existing", False),
                    extraction_confidence=item_data.get("extraction_confidence", 0.8),
                ))

        except Exception as e:
            error_msg = str(e)
            error_type = type(e).__name__
            logger.error("gpt4o_batch_failed", batch=batch_num, error=error_msg, error_type=error_type)

            # Check for quota/rate limit errors - these are fatal
            if "insufficient_quota" in error_msg.lower() or "rate_limit" in error_msg.lower():
                logger.error("openai_quota_exceeded", batch=batch_num)
                extraction_warnings.append(f"OpenAI quota exceeded - check billing")
                # Don't continue trying other batches if quota is exhausted
                break

            extraction_warnings.append(f"Batch {batch_num} failed: {error_msg}")
            continue

    logger.info("gpt4o_extraction_complete", total_items=len(all_items))

    return ExtractedBoQData(
        project_name=project_name,
        contractor_name=contractor_name,
        project_location=project_location,
        items=all_items,
        extraction_warnings=extraction_warnings,
    )


def _update_job_status_sync(
    supabase,
    job_id: str,
    status: Optional[BoQJobStatus] = None,
    progress: Optional[int] = None,
    error_message: Optional[str] = None,
) -> None:
    """Sync version of job status update."""
    update_data = {}
    if status:
        update_data["status"] = status.value
        if status == BoQJobStatus.PROCESSING:
            update_data["processing_started_at"] = datetime.utcnow().isoformat()
        elif status == BoQJobStatus.COMPLETED:
            update_data["completed_at"] = datetime.utcnow().isoformat()
    if progress is not None:
        update_data["progress_percent"] = progress
    if error_message:
        update_data["error_message"] = error_message
    if update_data:
        supabase.table("boq_jobs").update(update_data).eq("id", job_id).execute()


def _save_extracted_items_sync(supabase, job_id: str, extracted: ExtractedBoQData) -> None:
    """Sync version of saving extracted items."""
    if not extracted.items:
        return

    items_data = []
    for item in extracted.items:
        items_data.append({
            "job_id": job_id,
            "section": item.section,
            "item_number": item.item_number,
            "description": item.description,
            "unit": item.unit,
            "quantity": float(item.quantity) if item.quantity else None,
            "contractor_unit_price": float(item.contractor_unit_price) if item.contractor_unit_price else None,
            "contractor_total": float(item.contractor_total) if item.contractor_total else None,
            "item_type": item.item_type.value,
            "is_owner_supply": item.is_owner_supply,
            "is_existing": item.is_existing,
            "extraction_confidence": item.extraction_confidence,
        })

    # Insert in batches
    BATCH_SIZE = 50
    for i in range(0, len(items_data), BATCH_SIZE):
        batch = items_data[i:i + BATCH_SIZE]
        supabase.table("boq_items").insert(batch).execute()


def _calculate_summary_sync(supabase, job_id: str) -> None:
    """Sync version of calculating summary statistics."""
    result = supabase.table("boq_items").select("*").eq("job_id", job_id).execute()

    items = result.data if result.data else []

    materials_count = sum(1 for i in items if i.get("item_type") == "material")
    labor_count = sum(1 for i in items if i.get("item_type") == "labor")
    owner_supply_count = sum(1 for i in items if i.get("is_owner_supply"))

    contractor_total = sum(
        Decimal(str(i.get("contractor_total", 0) or 0))
        for i in items
    )

    supabase.table("boq_jobs").update({
        "materials_count": materials_count,
        "labor_count": labor_count,
        "owner_supply_count": owner_supply_count,
        "contractor_total": str(contractor_total),
        "market_estimate": "0.00",  # Placeholder
        "potential_savings": "0.00",  # Placeholder
    }).eq("id", job_id).execute()


def _extract_from_excel_sync(file_content: bytes, filename: str) -> ExtractedBoQData:
    """Synchronous Excel extraction."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl_not_installed")
        return ExtractedBoQData(
            extraction_warnings=["Excel parsing not available - openpyxl not installed"]
        )

    try:
        # Load workbook from bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active

        items = []
        current_section = None
        project_name = None
        contractor_name = None
        project_location = None

        # Scan for metadata in first 20 rows
        for row_idx in range(1, min(21, sheet.max_row + 1)):
            row_text = " ".join(
                str(cell.value or "").strip()
                for cell in sheet[row_idx]
                if cell.value
            ).lower()

            if "proyek" in row_text or "project" in row_text:
                for cell in sheet[row_idx]:
                    val = str(cell.value or "")
                    if ":" in val:
                        project_name = val.split(":", 1)[1].strip()
                        break
            elif "lokasi" in row_text or "location" in row_text:
                for cell in sheet[row_idx]:
                    val = str(cell.value or "")
                    if ":" in val:
                        project_location = val.split(":", 1)[1].strip()
                        break
            elif "cv" in row_text or "contractor" in row_text:
                for cell in sheet[row_idx]:
                    val = str(cell.value or "")
                    if val.lower().startswith("cv"):
                        contractor_name = val.strip()
                        break

        # Find header row
        header_row = None
        col_mapping = {}

        for row_idx in range(1, min(30, sheet.max_row + 1)):
            for col_idx, cell in enumerate(sheet[row_idx], 1):
                val = str(cell.value or "").upper().strip()
                if "URAIAN" in val or "DESCRIPTION" in val:
                    header_row = row_idx
                    for c_idx, c in enumerate(sheet[row_idx], 1):
                        c_val = str(c.value or "").upper().strip()
                        if "NO" == c_val or c_val == "#":
                            col_mapping["no"] = c_idx
                        elif "URAIAN" in c_val or "DESCRIPTION" in c_val:
                            col_mapping["description"] = c_idx
                        elif "SAT" in c_val or "UNIT" == c_val:
                            col_mapping["unit"] = c_idx
                        elif "VOL" in c_val or "QTY" in c_val or "QUANTITY" in c_val:
                            col_mapping["quantity"] = c_idx
                        elif "HARGA SATUAN" in c_val or "UNIT PRICE" in c_val:
                            col_mapping["unit_price"] = c_idx
                        elif "HARGA" == c_val or "TOTAL" == c_val or "AMOUNT" == c_val:
                            col_mapping["total"] = c_idx
                    break
            if header_row:
                break

        if not header_row or "description" not in col_mapping:
            return ExtractedBoQData(
                extraction_warnings=["Could not find BoQ table structure in Excel"]
            )

        # Extract items
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            desc_cell = sheet.cell(row_idx, col_mapping["description"])
            description = str(desc_cell.value or "").strip()

            if not description:
                continue

            # Check if section header
            if description.isupper() and len(description) > 5:
                qty_val = sheet.cell(row_idx, col_mapping.get("quantity", 1)).value
                if not qty_val:
                    current_section = description
                    continue

            unit = str(sheet.cell(row_idx, col_mapping.get("unit", 1)).value or "").strip()
            quantity = _parse_number(sheet.cell(row_idx, col_mapping.get("quantity", 1)).value)
            unit_price = _parse_number(sheet.cell(row_idx, col_mapping.get("unit_price", 1)).value)
            total = _parse_number(sheet.cell(row_idx, col_mapping.get("total", 1)).value)

            item_type = _classify_item(description)
            is_owner_supply = _check_owner_supply(description)
            is_existing = _check_existing(description)

            items.append(BoQItemExtracted(
                section=current_section,
                item_number=str(sheet.cell(row_idx, col_mapping.get("no", 1)).value or ""),
                description=description,
                unit=unit if unit else None,
                quantity=Decimal(str(quantity)) if quantity else None,
                contractor_unit_price=Decimal(str(unit_price)) if unit_price else None,
                contractor_total=Decimal(str(total)) if total else None,
                item_type=item_type,
                is_owner_supply=is_owner_supply,
                is_existing=is_existing,
                extraction_confidence=0.9,
            ))

        return ExtractedBoQData(
            project_name=project_name,
            contractor_name=contractor_name,
            project_location=project_location,
            items=items,
        )

    except Exception as e:
        logger.error("excel_extraction_failed", error=str(e))
        return ExtractedBoQData(
            extraction_warnings=[f"Excel extraction failed: {str(e)}"]
        )


async def _process_boq_job_async(
    job_id: str,
    file_content: bytes,
    file_format: BoQFileFormat,
    filename: str,
) -> None:
    """
    Async implementation of BoQ processing.
    """
    supabase = get_supabase_client()

    try:
        # Update status to processing
        await _update_job_status(
            supabase, job_id, BoQJobStatus.PROCESSING, progress=5
        )

        # Step 1: Extract items from file
        logger.info("boq_extraction_start", job_id=job_id, format=file_format.value)

        if file_format == BoQFileFormat.PDF:
            extracted = await _extract_from_pdf(file_content, filename)
        else:
            extracted = await _extract_from_excel(file_content, filename)

        await _update_job_status(supabase, job_id, progress=30)

        logger.info(
            "boq_extraction_complete",
            job_id=job_id,
            items_count=len(extracted.items),
        )

        # Step 2: Save extracted items and update job metadata
        await _save_extracted_items(supabase, job_id, extracted)

        # Update job with metadata
        supabase.table("boq_jobs").update({
            "project_name": extracted.project_name,
            "contractor_name": extracted.contractor_name,
            "project_location": extracted.project_location,
            "total_items_extracted": len(extracted.items),
        }).eq("id", job_id).execute()

        await _update_job_status(supabase, job_id, progress=40)

        # Step 3: Look up prices for material items
        logger.info("boq_pricing_start", job_id=job_id)
        await _lookup_material_prices(supabase, job_id)
        await _update_job_status(supabase, job_id, progress=85)

        # Step 4: Calculate summary statistics
        logger.info("boq_summary_start", job_id=job_id)
        await _calculate_summary(supabase, job_id)

        # Mark as completed
        await _update_job_status(
            supabase, job_id, BoQJobStatus.COMPLETED, progress=100
        )

        logger.info("boq_processing_complete", job_id=job_id)

    except Exception as e:
        logger.error("boq_processing_failed", job_id=job_id, error=str(e))
        await _update_job_status(
            supabase, job_id, BoQJobStatus.FAILED, error_message=str(e)
        )


# =============================================================================
# PDF Extraction (GPT-4o Vision)
# =============================================================================


async def _extract_from_pdf(file_content: bytes, filename: str) -> ExtractedBoQData:
    """Extract BoQ data from PDF using GPT-4o Vision.

    Converts PDF pages to images first since GPT-4o Vision only accepts image formats.
    Processes up to 10 pages to stay within token limits.
    """
    import fitz  # PyMuPDF
    import json
    from openai import AsyncOpenAI
    from app.config import get_settings

    settings = get_settings()
    # Use AsyncOpenAI for proper async/await support in background tasks
    client = AsyncOpenAI(
        api_key=settings.openai_api_key,
        timeout=180.0,  # 3 minute timeout for Vision API calls
    )

    # Convert PDF pages to images
    logger.info("pdf_to_images_start", filename=filename)

    try:
        pdf_document = fitz.open(stream=file_content, filetype="pdf")
        total_pages = len(pdf_document)

        # Process up to 10 pages to stay within token limits
        max_pages = min(total_pages, 10)
        logger.info("pdf_page_count", total_pages=total_pages, processing_pages=max_pages)

        # Convert pages to base64-encoded PNG images
        image_contents = []
        for page_num in range(max_pages):
            page = pdf_document[page_num]
            # Render at 150 DPI for good quality without excessive size
            mat = fitz.Matrix(150/72, 150/72)
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            base64_img = base64.b64encode(img_bytes).decode("utf-8")
            image_contents.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{base64_img}",
                    "detail": "high",
                },
            })

        pdf_document.close()
        logger.info("pdf_to_images_complete", image_count=len(image_contents))

    except Exception as e:
        logger.error("pdf_to_images_failed", error=str(e))
        return ExtractedBoQData(
            extraction_warnings=[f"PDF conversion failed: {str(e)}"]
        )

    extraction_prompt = """Analyze these pages from an Indonesian construction BoQ (Bill of Quantity / Rencana Anggaran Biaya) document.

Extract ALL line items into a structured JSON format. Pay attention to:
1. Section headers (e.g., "PEKERJAAN BONGKARAN", "PEKERJAAN KERAMIK")
2. Each line item with: description, unit (SAT), quantity (VOL), unit price (HARGA SATUAN), total (HARGA)
3. Items marked "(Suply By Owner)" or "(Supply By Owner)" - set is_owner_supply: true
4. Items marked "(use existing)" or "(existing)" - set is_existing: true

Classify each item:
- "material": Physical materials (granit, keramik, pipa, kabel, etc.)
- "labor": Work/service items (bongkar, instalasi, pasang, cat, plaster)
- "equipment": Installed equipment (pompa, AC unit, water heater)
- "unknown": Cannot determine

Return JSON in this exact format:
{
  "project_name": "extracted project name or null",
  "contractor_name": "contractor/CV name or null",
  "project_location": "location or null",
  "items": [
    {
      "section": "PEKERJAAN KERAMIK",
      "item_number": "F.1",
      "description": "Pas. Granit Lantai Master bedroom (Granit Suply By Owner)",
      "unit": "m2",
      "quantity": 19.47,
      "contractor_unit_price": 110000,
      "contractor_total": 2141700,
      "item_type": "material",
      "is_owner_supply": true,
      "is_existing": false,
      "extraction_confidence": 0.95
    }
  ],
  "extraction_warnings": ["any issues encountered"]
}

Be thorough - extract ALL items from ALL pages/sections. Indonesian terms:
- SAT = Satuan (Unit)
- VOL = Volume (Quantity)
- HARGA SATUAN = Unit Price
- HARGA = Total Price
- LS = Lump Sum
- m2 = square meters
- m1 = linear meters"""

    # Skip first page (usually cover with company info that gets refused)
    pages_to_process = image_contents[1:] if len(image_contents) > 1 else image_contents

    logger.info(
        "gpt4o_vision_extraction_start",
        total_pages=len(pages_to_process),
        skipped_cover=len(image_contents) > 1
    )

    # Process pages in small batches (3 pages at a time) for reliability
    # Large batches tend to timeout or get refused
    BATCH_SIZE = 3
    all_items = []
    extraction_warnings = []
    project_name = None
    contractor_name = None
    project_location = None

    for batch_start in range(0, len(pages_to_process), BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, len(pages_to_process))
        batch_pages = pages_to_process[batch_start:batch_end]
        batch_num = (batch_start // BATCH_SIZE) + 1
        total_batches = (len(pages_to_process) + BATCH_SIZE - 1) // BATCH_SIZE

        logger.info(
            "gpt4o_vision_batch_start",
            batch=batch_num,
            total_batches=total_batches,
            pages_in_batch=len(batch_pages),
        )

        try:
            content = [{"type": "text", "text": extraction_prompt}]
            content.extend(batch_pages)

            logger.info("gpt4o_api_call_starting", batch=batch_num, content_items=len(content))

            # Use native async OpenAI client - no thread needed
            response = await client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": content}],
                max_tokens=8000,
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            logger.info("gpt4o_api_call_returned", batch=batch_num)

            # Check for refusal
            choice = response.choices[0]
            refusal = getattr(choice.message, 'refusal', None)

            if refusal:
                logger.warning("gpt4o_vision_batch_refused", batch=batch_num, refusal=refusal)
                extraction_warnings.append(f"Batch {batch_num} refused: {refusal}")
                continue

            response_content = choice.message.content
            if not response_content:
                logger.warning("gpt4o_vision_batch_empty", batch=batch_num)
                extraction_warnings.append(f"Batch {batch_num} returned empty")
                continue

            data = json.loads(response_content)

            logger.info(
                "gpt4o_vision_batch_complete",
                batch=batch_num,
                items_count=len(data.get("items", [])),
                prompt_tokens=response.usage.prompt_tokens,
                completion_tokens=response.usage.completion_tokens,
            )

            # Extract metadata from first successful batch
            if not project_name:
                project_name = data.get("project_name")
            if not contractor_name:
                contractor_name = data.get("contractor_name")
            if not project_location:
                project_location = data.get("project_location")

            # Accumulate items
            for item_data in data.get("items", []):
                all_items.append(BoQItemExtracted(
                    section=item_data.get("section"),
                    item_number=item_data.get("item_number"),
                    description=item_data.get("description", ""),
                    unit=item_data.get("unit"),
                    quantity=item_data.get("quantity"),
                    contractor_unit_price=item_data.get("contractor_unit_price"),
                    contractor_total=item_data.get("contractor_total"),
                    item_type=BoQItemType(item_data.get("item_type", "unknown")),
                    is_owner_supply=item_data.get("is_owner_supply", False),
                    is_existing=item_data.get("is_existing", False),
                    extraction_confidence=item_data.get("extraction_confidence", 0.8),
                ))

            # Add any warnings from this batch
            extraction_warnings.extend(data.get("extraction_warnings", []))

        except Exception as e:
            logger.error("gpt4o_vision_batch_failed", batch=batch_num, error=str(e))
            extraction_warnings.append(f"Batch {batch_num} failed: {str(e)}")
            continue

    logger.info("gpt4o_vision_extraction_complete", total_items=len(all_items))

    return ExtractedBoQData(
        project_name=project_name,
        contractor_name=contractor_name,
        project_location=project_location,
        items=all_items,
        extraction_warnings=extraction_warnings,
    )


async def _extract_pages_individually(
    client, image_contents: list, prompt: str, logger
) -> ExtractedBoQData:
    """Fall back to extracting pages one by one if batch extraction fails."""
    import json

    all_items = []
    warnings = []

    for idx, img_content in enumerate(image_contents):
        try:
            response = await asyncio.to_thread(
                lambda img=img_content: client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{
                        "role": "user",
                        "content": [{"type": "text", "text": prompt}, img]
                    }],
                    max_tokens=4000,
                    temperature=0.1,
                    response_format={"type": "json_object"},
                )
            )

            choice = response.choices[0]
            refusal = getattr(choice.message, 'refusal', None)

            if refusal:
                logger.warning("page_refused", page=idx, refusal=refusal)
                warnings.append(f"Page {idx} refused: {refusal}")
                continue

            if not choice.message.content:
                continue

            data = json.loads(choice.message.content)
            for item_data in data.get("items", []):
                all_items.append(BoQItemExtracted(
                    section=item_data.get("section"),
                    item_number=item_data.get("item_number"),
                    description=item_data.get("description", ""),
                    unit=item_data.get("unit"),
                    quantity=item_data.get("quantity"),
                    contractor_unit_price=item_data.get("contractor_unit_price"),
                    contractor_total=item_data.get("contractor_total"),
                    item_type=BoQItemType(item_data.get("item_type", "unknown")),
                    is_owner_supply=item_data.get("is_owner_supply", False),
                    is_existing=item_data.get("is_existing", False),
                    extraction_confidence=item_data.get("extraction_confidence", 0.8),
                ))

            logger.info("page_extracted", page=idx, items=len(data.get("items", [])))

        except Exception as e:
            logger.warning("page_extraction_failed", page=idx, error=str(e))
            warnings.append(f"Page {idx} failed: {str(e)}")

    return ExtractedBoQData(
        items=all_items,
        extraction_warnings=warnings if warnings else None,
    )


# =============================================================================
# Excel Extraction
# =============================================================================


async def _extract_from_excel(file_content: bytes, filename: str) -> ExtractedBoQData:
    """Extract BoQ data from Excel file."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl_not_installed")
        return ExtractedBoQData(
            extraction_warnings=["Excel parsing not available - openpyxl not installed"]
        )

    try:
        # Load workbook from bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active

        items = []
        current_section = None
        project_name = None
        contractor_name = None
        project_location = None

        # Scan for metadata in first 20 rows
        for row_idx in range(1, min(21, sheet.max_row + 1)):
            row_text = " ".join(
                str(cell.value or "").strip()
                for cell in sheet[row_idx]
                if cell.value
            ).lower()

            if "proyek" in row_text or "project" in row_text:
                # Try to extract project name
                for cell in sheet[row_idx]:
                    val = str(cell.value or "")
                    if ":" in val:
                        project_name = val.split(":", 1)[1].strip()
                        break
            elif "lokasi" in row_text or "location" in row_text:
                for cell in sheet[row_idx]:
                    val = str(cell.value or "")
                    if ":" in val:
                        project_location = val.split(":", 1)[1].strip()
                        break
            elif "cv" in row_text or "contractor" in row_text:
                for cell in sheet[row_idx]:
                    val = str(cell.value or "")
                    if val.lower().startswith("cv"):
                        contractor_name = val.strip()
                        break

        # Find header row (look for "URAIAN" or "DESCRIPTION")
        header_row = None
        col_mapping = {}

        for row_idx in range(1, min(30, sheet.max_row + 1)):
            for col_idx, cell in enumerate(sheet[row_idx], 1):
                val = str(cell.value or "").upper().strip()
                if "URAIAN" in val or "DESCRIPTION" in val:
                    header_row = row_idx
                    # Map columns
                    for c_idx, c in enumerate(sheet[row_idx], 1):
                        c_val = str(c.value or "").upper().strip()
                        if "NO" == c_val or c_val == "#":
                            col_mapping["no"] = c_idx
                        elif "URAIAN" in c_val or "DESCRIPTION" in c_val:
                            col_mapping["description"] = c_idx
                        elif "SAT" in c_val or "UNIT" == c_val:
                            col_mapping["unit"] = c_idx
                        elif "VOL" in c_val or "QTY" in c_val or "QUANTITY" in c_val:
                            col_mapping["quantity"] = c_idx
                        elif "HARGA SATUAN" in c_val or "UNIT PRICE" in c_val:
                            col_mapping["unit_price"] = c_idx
                        elif "HARGA" == c_val or "TOTAL" == c_val or "AMOUNT" == c_val:
                            col_mapping["total"] = c_idx
                    break
            if header_row:
                break

        if not header_row or "description" not in col_mapping:
            return ExtractedBoQData(
                extraction_warnings=["Could not find BoQ table structure in Excel"]
            )

        # Extract items
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            desc_cell = sheet.cell(row_idx, col_mapping["description"])
            description = str(desc_cell.value or "").strip()

            if not description:
                continue

            # Check if this is a section header
            if description.isupper() and len(description) > 5:
                # Check if it looks like a section (no quantity/price)
                qty_val = sheet.cell(row_idx, col_mapping.get("quantity", 1)).value
                if not qty_val:
                    current_section = description
                    continue

            # Extract item data
            unit = str(sheet.cell(row_idx, col_mapping.get("unit", 1)).value or "").strip()
            quantity = _parse_number(sheet.cell(row_idx, col_mapping.get("quantity", 1)).value)
            unit_price = _parse_number(sheet.cell(row_idx, col_mapping.get("unit_price", 1)).value)
            total = _parse_number(sheet.cell(row_idx, col_mapping.get("total", 1)).value)

            # Classify item
            item_type = _classify_item(description)
            is_owner_supply = _check_owner_supply(description)
            is_existing = _check_existing(description)

            items.append(BoQItemExtracted(
                section=current_section,
                item_number=str(sheet.cell(row_idx, col_mapping.get("no", 1)).value or ""),
                description=description,
                unit=unit if unit else None,
                quantity=Decimal(str(quantity)) if quantity else None,
                contractor_unit_price=Decimal(str(unit_price)) if unit_price else None,
                contractor_total=Decimal(str(total)) if total else None,
                item_type=item_type,
                is_owner_supply=is_owner_supply,
                is_existing=is_existing,
                extraction_confidence=0.9,  # Excel is more reliable
            ))

        return ExtractedBoQData(
            project_name=project_name,
            contractor_name=contractor_name,
            project_location=project_location,
            items=items,
        )

    except Exception as e:
        logger.error("excel_extraction_failed", error=str(e))
        return ExtractedBoQData(
            extraction_warnings=[f"Excel extraction failed: {str(e)}"]
        )


# =============================================================================
# Helper Functions
# =============================================================================


def _parse_number(value) -> Optional[float]:
    """Parse a number from cell value."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        # Remove thousand separators and convert
        cleaned = str(value).replace(",", "").replace(".", "").strip()
        if cleaned:
            return float(cleaned)
    except ValueError:
        pass
    return None


def _classify_item(description: str) -> BoQItemType:
    """Classify an item as material, labor, or equipment.

    Priority: Labor prefixes (bongkar, instalasi) > Material names > Equipment
    This ensures "bongkar pintu" is labor, not material.
    """
    desc_lower = description.lower()

    # Labor-action prefixes that override material indicators
    # These indicate the WORK of doing something to a material
    labor_action_prefixes = [
        "bongkar",       # demolition/removal
        "instalasi",     # installation (labor)
        "pasang ",       # mounting (labor, not Pas. for tile)
        "pek.",          # pekerjaan (work)
        "pek ",
        "perbaikan",     # repair
        "pengecatan",    # painting work
        "pembuangan",    # disposal
        "cleaning",      # cleaning
    ]

    # Check for labor-action prefixes first (highest priority)
    for prefix in labor_action_prefixes:
        if desc_lower.startswith(prefix) or f" {prefix}" in desc_lower[:30]:
            return BoQItemType.LABOR

    # Check for equipment (standalone items, not installation)
    equipment_terms = ["pompa", "ac unit", "water heater", "filter kolam"]
    for term in equipment_terms:
        if term in desc_lower and "instalasi" not in desc_lower:
            return BoQItemType.EQUIPMENT

    # Check for material indicators
    for indicator in MATERIAL_INDICATORS:
        if indicator in desc_lower:
            return BoQItemType.MATERIAL

    # Check remaining labor indicators
    for indicator in LABOR_INDICATORS:
        if indicator in desc_lower:
            return BoQItemType.LABOR

    return BoQItemType.UNKNOWN


def _check_owner_supply(description: str) -> bool:
    """Check if item is marked as 'Supply By Owner'."""
    desc_lower = description.lower()
    for pattern in OWNER_SUPPLY_PATTERNS:
        if re.search(pattern, desc_lower):
            return True
    return False


def _check_existing(description: str) -> bool:
    """Check if item is marked as 'use existing'."""
    desc_lower = description.lower()
    for pattern in EXISTING_PATTERNS:
        if re.search(pattern, desc_lower):
            return True
    return False


def _normalize_material_name(description: str) -> str:
    """
    Normalize a material description for Tokopedia search.

    Removes installation prefixes, owner supply notes, etc.
    """
    # Remove common prefixes
    prefixes_to_remove = [
        r"^pas\.\s*",      # Pasang
        r"^pas\s+",
        r"^instalasi\s+",
        r"^pek\.\s*",
        r"^pek\s+",
    ]

    result = description.lower()
    for prefix in prefixes_to_remove:
        result = re.sub(prefix, "", result, flags=re.IGNORECASE)

    # Remove owner supply / existing notes (with or without parentheses)
    # Pattern handles: "(Granit Suply By Owner)", "Suply By Owner", "(unit supply by owner)"
    result = re.sub(r"\([^)]*suply\s*by\s*owner[^)]*\)", "", result, flags=re.IGNORECASE)
    result = re.sub(r"\([^)]*supply\s*by\s*owner[^)]*\)", "", result, flags=re.IGNORECASE)
    result = re.sub(r"\(?use\s*existing\)?", "", result, flags=re.IGNORECASE)
    result = re.sub(r"\([^)]*existing[^)]*\)", "", result, flags=re.IGNORECASE)

    # Remove location/room specifiers
    result = re.sub(r"master\s*bed\s*room", "", result, flags=re.IGNORECASE)
    result = re.sub(r"master\s*bathroom", "", result, flags=re.IGNORECASE)
    result = re.sub(r"living\s*dining\s*kitchen", "", result, flags=re.IGNORECASE)
    result = re.sub(r"lantai\s*\d+", "", result, flags=re.IGNORECASE)
    result = re.sub(r"area\s+\w+", "", result, flags=re.IGNORECASE)

    # Clean up
    result = re.sub(r"\s+", " ", result).strip()

    return result


async def _update_job_status(
    supabase,
    job_id: str,
    status: Optional[BoQJobStatus] = None,
    progress: Optional[int] = None,
    error_message: Optional[str] = None,
) -> None:
    """Update job status in database."""
    update_data = {}

    if status:
        update_data["status"] = status.value
        if status == BoQJobStatus.PROCESSING:
            update_data["processing_started_at"] = datetime.utcnow().isoformat()
        elif status == BoQJobStatus.COMPLETED:
            update_data["completed_at"] = datetime.utcnow().isoformat()

    if progress is not None:
        update_data["progress_percent"] = progress

    if error_message:
        update_data["error_message"] = error_message

    if update_data:
        supabase.table("boq_jobs").update(update_data).eq("id", job_id).execute()


async def _save_extracted_items(
    supabase,
    job_id: str,
    extracted: ExtractedBoQData,
) -> None:
    """Save extracted items to database."""
    if not extracted.items:
        return

    items_data = []
    for item in extracted.items:
        items_data.append({
            "job_id": job_id,
            "section": item.section,
            "item_number": item.item_number,
            "description": item.description,
            "unit": item.unit,
            "quantity": float(item.quantity) if item.quantity else None,
            "contractor_unit_price": float(item.contractor_unit_price) if item.contractor_unit_price else None,
            "contractor_total": float(item.contractor_total) if item.contractor_total else None,
            "item_type": item.item_type.value,
            "is_owner_supply": item.is_owner_supply,
            "is_existing": item.is_existing,
            "extraction_confidence": item.extraction_confidence,
        })

    # Batch insert
    supabase.table("boq_items").insert(items_data).execute()


async def _lookup_material_prices(supabase, job_id: str) -> None:
    """Look up Tokopedia prices for material items."""
    # Get material items
    result = (
        supabase.table("boq_items")
        .select("*")
        .eq("job_id", job_id)
        .eq("item_type", "material")
        .execute()
    )

    items = result.data or []
    total_materials = len(items)

    for idx, item in enumerate(items):
        try:
            # Normalize material name for search
            search_query = _normalize_material_name(item["description"])

            if len(search_query) < 3:
                continue  # Skip very short queries

            # Look up price
            price_result = await get_best_material_price(search_query)

            # Calculate confidence based on how well name matches
            match_confidence = 0.0
            if price_result.get("name"):
                # Simple word overlap scoring
                search_words = set(search_query.lower().split())
                result_words = set(price_result["name"].lower().split())
                overlap = len(search_words & result_words)
                total_words = len(search_words | result_words)
                match_confidence = overlap / total_words if total_words > 0 else 0

            # Calculate price comparison
            market_price = price_result.get("price_idr", 0)
            contractor_price = item.get("contractor_unit_price", 0) or 0

            price_diff = None
            price_diff_percent = None
            if market_price > 0 and contractor_price > 0:
                price_diff = float(contractor_price) - market_price
                price_diff_percent = (price_diff / market_price) * 100

            # Calculate market total
            quantity = item.get("quantity", 0) or 0
            market_total = market_price * float(quantity) if market_price else None

            # Update item
            update_data = {
                "search_query": search_query,
                "tokopedia_product_name": price_result.get("name"),
                "tokopedia_price": market_price if market_price else None,
                "tokopedia_url": price_result.get("url"),
                "tokopedia_seller": price_result.get("seller"),
                "tokopedia_seller_location": price_result.get("seller_location"),
                "tokopedia_rating": price_result.get("rating"),
                "tokopedia_sold_count": price_result.get("sold_count"),
                "match_confidence": round(match_confidence, 3),
                "market_unit_price": market_price if market_price else None,
                "market_total": market_total,
                "price_difference": price_diff,
                "price_difference_percent": round(price_diff_percent, 2) if price_diff_percent else None,
            }

            supabase.table("boq_items").update(update_data).eq("id", item["id"]).execute()

            # Update progress
            progress = 40 + int((idx + 1) / total_materials * 45)  # 40% to 85%
            await _update_job_status(supabase, job_id, progress=progress)

            # Small delay to avoid rate limiting
            await asyncio.sleep(0.2)

        except Exception as e:
            logger.warning(
                "price_lookup_failed",
                item_id=item["id"],
                description=item["description"],
                error=str(e),
            )
            continue


async def _calculate_summary(supabase, job_id: str) -> None:
    """Calculate summary statistics for the job."""
    # Get all items
    result = (
        supabase.table("boq_items")
        .select("*")
        .eq("job_id", job_id)
        .execute()
    )

    items = result.data or []

    # Calculate totals
    contractor_total = sum(
        float(item.get("contractor_total") or 0)
        for item in items
    )

    market_estimate = sum(
        float(item.get("market_total") or 0)
        for item in items
        if item.get("item_type") == "material" and item.get("market_total")
    )

    # Count by type
    materials_count = sum(1 for item in items if item.get("item_type") == "material")
    labor_count = sum(1 for item in items if item.get("item_type") == "labor")
    owner_supply_count = sum(1 for item in items if item.get("is_owner_supply"))

    # Calculate savings (only for materials we could price)
    potential_savings = max(0, contractor_total - market_estimate) if market_estimate > 0 else 0

    # Update job
    supabase.table("boq_jobs").update({
        "contractor_total": contractor_total,
        "market_estimate": market_estimate,
        "potential_savings": potential_savings,
        "materials_count": materials_count,
        "labor_count": labor_count,
        "owner_supply_count": owner_supply_count,
    }).eq("id", job_id).execute()
